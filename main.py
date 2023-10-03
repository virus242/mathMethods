from mathMethods.methods import MathMethod
import config.config as config
import openpyxl
import copy
import telebot
from telebot import types

#AArr = [100, 250, 200, 300]
#BArr = [200, 200, 100, 100, 250]

#AArr = [15, 9, 13, 8, 9]
#BArr = [21, 15, 7, 6]


#CArr = [[10, 7, 4, 1, 4],
#     [2, 7, 10, 6, 11],
#     [8, 5, 3, 2, 2],
#     [11, 8, 12, 16, 13]] 

#CArr = [[2, 6, 5, 3],
#     [3, 2, 1, 4],
#     [3, 6, 2, 5],
#     [3, 6, 5, 6],
#     [3, 6, 5, 7]] 
obj = MathMethod()
FlagOnSaveDetails = []

bot = telebot.TeleBot(config.TOKEN)
@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, 
                     "Для того чтобы использовать бота необходимо отправить ему файл в формате xlsx. Файл должен быть оформлен в следующем виде." +
                     "Если вы хотите включить подробный разбор введите комманду '/include_details', а если вы хотите их выключить введите '/disable_details'")
    photo = open("Example.png", 'rb')
    bot.send_photo(message.chat.id, photo)

@bot.message_handler(commands=['include_details'])
def include_details(message):
    bot.send_message(message.chat.id, "Теперь подробности будут видны")
    obj.setFlagOnSaveDetails(True)
    print(obj.FlagOnSaveDetails)

@bot.message_handler(commands=['disable_details'])
def disable_details(message):
    bot.send_message(message.chat.id, "Убрал подробности")
    obj.setFlagOnSaveDetails(False)
    print(obj.FlagOnSaveDetails)

@bot.message_handler(content_types=['document'])
def handle_docs_photo(message):
    print("in if Flag : ", f"{obj.getFlagOnSaveDetails()}")
    chat_id = message.chat.id

    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)

    src = message.document.file_name;
    with open(src, 'wb') as new_file:
        new_file.write(downloaded_file)

    bot.reply_to(message, "Работаем...")


    AArr = []
    BArr = []

    CArr = []
    data = openpyxl.load_workbook("task.xlsx").active

    for i in range(0, data.max_row):
        CArr.append([])
        countColumn = 0
        for col in data.iter_cols(1, data.max_column):
            
            if countColumn == data.max_column-1:
                AArr.append(col[i].value)
            elif i == data.max_row-1:
                BArr.append(col[i].value)
            else:
                CArr[i].append(col[i].value)
            countColumn+=1    

    AArr = AArr[:-1]    #delete None

    
    
    obj.setArgs(copy.deepcopy(AArr), copy.deepcopy(BArr), copy.deepcopy(CArr))
    northwestMethodRes = obj.northwestMethod() 
    print("in if Flag : ", f"{obj.getFlagOnSaveDetails()}")
    if obj.FlagOnSaveDetails:
        print("send")
        bot.send_message(chat_id, f"F = {obj.getDetails()[:-1]} = {northwestMethodRes}")
    bot.send_message(chat_id, "Метод северо-западного угла = " + str(northwestMethodRes))
    
   

    obj.setArgs(copy.deepcopy(AArr), copy.deepcopy(BArr), copy.deepcopy(CArr))
    minCostMethodRes = obj.minCostMethod()
    if obj.FlagOnSaveDetails:
        bot.send_message(chat_id, f"F = {obj.getDetails()[:-1]} = {minCostMethodRes}")
    bot.send_message(chat_id, "Метод минимального остатка = " + str(minCostMethodRes))



bot.polling(none_stop=True, interval=0)

